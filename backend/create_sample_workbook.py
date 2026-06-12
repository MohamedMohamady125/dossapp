"""Generate a sample Excel workbook for testing — ~20 athletes with varied data."""

from openpyxl import Workbook
from datetime import datetime

wb = Workbook()

# ===== Sheet 1: June Reg. (roster) =====
ws = wb.active
ws.title = " June Reg."

headers = [
    "f", "M Code", "ID", "Name", "Age", "Date of Birth", "gender",
    "Step", "Season", "Type", "Days ", "Times", "Sessions",
    "pay ", "Phone 1", "Phone 2", "Receipt No.", "visa", "Comment"
]
ws.append(headers)

# Excel serial date for 2015-03-15 = 42078
# Excel serial date for 2010-07-22 = 40381
# Excel serial date for 2018-01-05 = 43105
athletes = [
    ["Gems Sch.", "", "M1", "Ahmed Hassan", "", 42078.0, "M", "St.1", "June", "Class", "Sat.Wed", "", "8 Sessions", "1200", "01012345678", "01198765432", "", "", ""],
    ["Outsider", "", "M2", "Sara Mohamed", "", 40381.0, "F", "St.3", "June", "Class", "Sun.Tues", "", "8 Sessions", "1500", "01234567890", "", "", "", "50% D"],
    ["Gems Sch.", "", "M3", "Omar Ali", "", 43105.0, "M", "P.p1", "June", "Class", "Sat.Wed", "", "8 Sessions", "1200", "2.01E+11", "", "", "", ""],
    ["", "", "M4", "Nour Ibrahim", "", 41456.0, "F", "St.2", "June", "Private.1", "Mon.Thu", "", "4 Sessions", "2500", "0100 555 1234", "", "", "", ""],
    ["Outsider", "", "M5", "Youssef Khaled", "", 42890.0, "M", "St.5", "June", "Semi-Private.2", "Sat.Wed", "", "8 Sessions", "1800", "+201112223334", "", "R-001", "", ""],
    ["Gems Sch.", "", "M6", "Layla Mostafa", "", 43500.0, "F", "St.1", "June", "Class", "Sun.Tues", "", "8 Sessions", "1200", "01001112233", "01009998877", "", "", "20%Dec"],
    ["", "", "M7", "Karim Adel", "", None, "M", "Pre Team", "June", "Pre Team", "Sat.Wed", "", "12 Sessions", "2000", "01155544433", "", "", "", ""],
    ["Outsider", "", "M8", "Malak Samir", "", 41000.0, "F", "St.4", "June", "Class", "Sun.Tues", "", "8 Sessions", "1500", "01066677788", "", "R-002", "", ""],
    ["Gems Sch.", "", "M9", "Ziad Tarek", "", 43800.0, "M", "", "June", "Class", "Sat.Wed", "", "8 Sessions", "", "01177788899", "", "", "", "ترحيل"],
    ["", "", "M10", "Hana Essam", "", 40200.0, "F", "Junior Team", "June", "Junior Teams", "Mon.Thu", "", "12 Sessions", "1800", "\\01022334455", "", "", "", ""],
    ["Outsider", "", "M11", "Tamer Fathy", "", 42500.0, "M", "St.6", "June", "Class", "Sat.Wed", "", "8 Sessions", "1500", "01033445566", "", "", "", ""],
    ["Gems Sch.", "", "M12", "Dina Waleed", "", 43200.0, "F", "St.2", "June", "Class", "Sun.Tues", "", "8 Sessions", "1200", "01044556677", "", "R-003", "", ""],
    ["", "", "M13", "Ali Mansour", "", 41800.0, "M", "St.7", "June", "Semi-Private.2", "Sat.Wed", "", "8 Sessions", "1800", "01055667788", "01099887766", "", "", ""],
    ["Outsider", "", "M14", "Rana Hossam", "", 42000.0, "", "St.3", "June", "Class", "Sun.Tues", "", "8 Sessions", "1500", "01066778899", "", "", "", ""],
    ["Gems Sch.", "", "M15", "Mahmoud Gamal", "", 43600.0, "M", "Adult Men", "June", "Class", "Mon.Thu", "", "8 Sessions", "1200", "01077889900", "", "", "", ""],
    ["", "", "M16", "Fatma Nabil", "", 40600.0, "F", "Adult Women", "June", "Class", "Sat.Wed", "", "8 Sessions", "1500", "01088990011", "", "", "", ""],
    ["Gems Sch.", "", "M17", "Hussein Reda", "", 44000.0, "M", "St.1", "June", "Class", "Sun.Tues", "", "4 Sessions", "700", "0109 901 0122", "", "", "", ""],
    ["Outsider", "", "M18", "Salma Ashraf", "", 43400.0, "F", "St.9", "June", "Class", "Sat.Wed", "", "8 Sessions", "1500", " 01100112233 ", "", "", "", ""],
    # Missing DOB row (should compute garbage age — we skip it)
    ["", "", "M19", "Amr Sabry", "", "", "M", "St.4", "June", "Private.1", "Sun.Tues", "", "4 Sessions", "2500", "01111223344", "", "", "", ""],
    # Multi-slot athlete (appears in two attendance sheets)
    ["Gems Sch.", "", "M20", "Yasmin Sherif", "", 43100.0, "F", "St.5", "June", "Class", "Sat.Wed", "", "8 Sessions", "1200", "01122334455", "", "", "", ""],
]

for row in athletes:
    ws.append(row)


# ===== Sheet 2: Sat.Wed (attendance) =====
ws2 = wb.create_sheet("Sat.Wed")
ws2.append(["", "", "", "Saturday - Wednesday Schedule"])
ws2.append([])
ws2.append(["", "4:15 to 5:10 pm", "", ""])  # time block
ws2.append(["Coach Ahmed", "", "", ""])  # coach
ws2.append(["M1", "Ahmed Hassan", "Present", ""])
ws2.append(["M3", "Omar Ali", "Present", ""])
ws2.append(["M5", "Youssef Khaled", "Absent", ""])
ws2.append([])
ws2.append(["Coach Sara", "", "", ""])  # another coach
ws2.append(["M11", "Tamer Fathy", "Present", ""])
ws2.append(["M13", "Ali Mansour", "Present", ""])
ws2.append(["M20", "Yasmin Sherif", "Present", ""])
ws2.append([])
ws2.append(["", "5:15 to 6:10 pm", "", ""])  # another time block
ws2.append(["Coach Mohamed", "", "", ""])
ws2.append(["M16", "Fatma Nabil", "Present", ""])
ws2.append(["M18", "Salma Ashraf", "Present", ""])


# ===== Sheet 3: Sun.Tues (attendance) =====
ws3 = wb.create_sheet("Sun.Tues")
ws3.append(["", "", "", "Sunday - Tuesday Schedule"])
ws3.append([])
ws3.append(["", "4:15 to 5:10 pm", "", ""])
ws3.append(["Coach Laila", "", "", ""])
ws3.append(["M2", "Sara Mohamed", "Present", ""])
ws3.append(["M6", "Layla Mostafa", "Present", ""])
ws3.append(["M8", "Malak Samir", "Present", ""])
ws3.append([])
ws3.append(["Coach Amr", "", "", ""])
ws3.append(["M14", "Rana Hossam", "Present", ""])
ws3.append(["M17", "Hussein Reda", "Present", ""])
ws3.append(["M20", "Yasmin Sherif", "Present", ""])  # Multi-slot: also in Sat.Wed
ws3.append([])
ws3.append(["", "5:15 to 6:10 pm", "", ""])
ws3.append(["Coach Dalia", "", "", ""])
ws3.append(["M4", "Nour Ibrahim", "Present", ""])
ws3.append(["M19", "Amr Sabry", "Present", ""])


# ===== Sheet 4: Mon.Thu (empty template) =====
ws4 = wb.create_sheet("Mon.Thu")
ws4.append(["", "", "", "Monday - Thursday Schedule"])
ws4.append(["", "No schedule data for this period"])


# ===== Sheet 5: Teams (empty template) =====
ws5 = wb.create_sheet("Teams")
ws5.append(["", "", "", "Teams - May Template"])


# ===== Sheet 6: Sk. (price matrix + coach list) =====
ws6 = wb.create_sheet("Sk.")
# Price matrix headers
ws6.append(["Type", "8x", "4x", "2x", "1x", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "Coaches"])
ws6.append(["Class (Regular)", "1500", "900", "500", "300", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "Coach Ahmed"])
ws6.append(["Class (GEMS)", "1200", "700", "400", "250", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "Coach Sara"])
ws6.append(["Private 1:1", "2500", "1400", "800", "450", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "Coach Mohamed"])
ws6.append(["Semi Priv 1:2", "1800", "1000", "600", "350", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "Coach Laila"])
ws6.append(["Semi Priv 1:3", "1600", "900", "550", "300", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "Coach Amr"])
ws6.append(["Junior Teams", "1800", "1000", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "Coach Dalia"])
ws6.append(["Elite Team", "2200", "1200", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "Coach Khaled"])

output_path = "tests/sample_workbook.xlsx"
wb.save(output_path)
print(f"Sample workbook saved to {output_path}")
print(f"Athletes: {len(athletes)}")
