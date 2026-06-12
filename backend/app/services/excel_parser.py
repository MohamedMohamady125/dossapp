"""Parse Aqua Athletic Excel workbooks into Athlete objects.

Universal parser — handles all branch format variations:
- Rehab/Madinaty: 19-col layout with M Code + visa, M-prefix IDs, M/F gender
- Choueifat/Choueifat Cairo: 17-col layout, C/E-prefix IDs, Male/Female gender
- Different header names: pay/Pay/Fees/fees, Date of Birth/Birth Date/B . D, etc.
- Attendance sheets with coach in Col1, repeated sub-headers, Pay in attendance
"""

import logging
import re
from dataclasses import dataclass
from datetime import datetime, timedelta
from typing import Optional

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.services.roster_source import Athlete, ScheduleEntry

logger = logging.getLogger(__name__)

EXCEL_EPOCH = datetime(1899, 12, 30)

# Map of lowercase/stripped header → field name (all known variations)
HEADER_MAP = {
    "f": "segment", "membership": "segment", "member": "segment",
    "id": "id",
    "name": "name",
    "date of birth": "dob", "birth date": "dob", "dob": "dob", "b . d": "dob", "b.d": "dob",
    "gender": "gender",
    "step": "step", "level": "step", "st.": "step",
    "type": "type",
    "days": "days",
    "times": "times",
    "season": "season",
    "sessions": "sessions",
    "pay": "pay", "fees": "pay", "amount": "pay",
    "phone 1": "phone1", "phone": "phone1", "phone1": "phone1",
    "phone 2": "phone2", "phone2": "phone2",
    "receipt no.": "receipt_no", "receipt num": "receipt_no",
    "receipt number": "receipt_no", "receipt no": "receipt_no",
    "comment": "comment", "comments": "comment", "notes": "comment",
    "visa": "visa",
    "knowing method": "knowing_method",
    "m code": "m_code",
}


def _clean(val) -> str:
    if val is None:
        return ""
    s = str(val).strip()
    if s in ("#REF!", "#N/A", "#VALUE!", "#DIV/0!"):
        return ""
    return s


def _parse_athlete_number(id_val: str) -> Optional[int]:
    """Extract number from ID like 'M31', 'C1001', 'E1002' → number."""
    if not id_val:
        return None
    nums = re.findall(r"\d+", id_val)
    if nums:
        return int(nums[0])
    return None


def _parse_dob(raw) -> Optional[str]:
    """Convert various DOB formats to ISO date string."""
    if raw is None:
        return None
    if isinstance(raw, datetime):
        return raw.strftime("%Y-%m-%d")
    s = str(raw).strip()
    if not s:
        return None
    # Try Excel serial number
    try:
        num = float(s)
        if 1 < num < 100000:
            dt = EXCEL_EPOCH + timedelta(days=num)
            return dt.strftime("%Y-%m-%d")
    except (ValueError, TypeError, OverflowError):
        pass
    # Try various string date formats
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d/%m/%y", "%d\\%m\\%Y",
                "%d\\%m\\%y", "%m/%d/%Y", "%m/%d/%y"):
        try:
            dt = datetime.strptime(s, fmt)
            if dt.year > 2050:
                dt = dt.replace(year=dt.year - 100)
            return dt.strftime("%Y-%m-%d")
        except ValueError:
            pass
    # Handle weird separators like 'L', '.'
    cleaned = re.sub(r"[LlBb.\\/-]", "/", s)
    for fmt in ("%d/%m/%Y", "%d/%m/%y"):
        try:
            dt = datetime.strptime(cleaned, fmt)
            if dt.year > 2050:
                dt = dt.replace(year=dt.year - 100)
            return dt.strftime("%Y-%m-%d")
        except ValueError:
            pass
    return None


def _compute_age(dob_str: Optional[str]) -> Optional[float]:
    if not dob_str:
        return None
    try:
        dob = datetime.strptime(dob_str, "%Y-%m-%d")
        age = (datetime.now() - dob).days / 365.25
        if age < 0 or age > 100:
            return None
        return round(age, 1)
    except (ValueError, TypeError):
        return None


def _normalize_gender(raw: str) -> Optional[str]:
    if not raw:
        return None
    g = raw.strip().lower()
    if g in ("m", "male", "boy", "ذكر"):
        return "M"
    if g in ("f", "female", "girl", "أنثى"):
        return "F"
    return raw.strip() or None


def _clean_pay(raw: str) -> Optional[str]:
    """Clean pay value — return numeric string or None."""
    if not raw:
        return None
    s = raw.strip().lower()
    # Skip non-numeric values like 'Trans', 'Pay', header repeats
    if s in ("pay", "fees", "trans", "trans ", "0", "0.0", ""):
        return None
    cleaned = re.sub(r"[^\d.]", "", s)
    if not cleaned:
        return None
    try:
        float(cleaned)
        return cleaned
    except ValueError:
        return None


def _find_header_row(ws: Worksheet) -> tuple[Optional[int], dict[str, int]]:
    """Find the header row and map field names to column indices."""
    max_row = min(10, ws.max_row or 10)
    max_col = min(50, ws.max_column or 50)

    for row_idx in range(1, max_row + 1):
        cells = {}
        for col_idx in range(1, max_col + 1):
            val = _clean(ws.cell(row=row_idx, column=col_idx).value).lower().strip()
            if val:
                cells[val] = col_idx

        # Match known headers
        matches = {}
        for header_text, field_name in HEADER_MAP.items():
            if header_text in cells and field_name not in matches:
                matches[field_name] = cells[header_text]

        # Need at least ID + Name to consider this a valid header row
        if "id" in matches and "name" in matches:
            return row_idx, matches

    # Fallback: try to detect by column position if no clear headers
    return None, {}


def _detect_gender_column(ws: Worksheet, header_row: int, col_map: dict) -> Optional[int]:
    """Some branches (Choueifat Cairo) have no gender header — detect by values."""
    if "gender" in col_map:
        return col_map["gender"]
    # Check columns near where gender typically is (5-8)
    for col in range(5, 9):
        if col in col_map.values():
            continue
        # Check first 5 data rows
        vals = set()
        for row in range(header_row + 1, min(header_row + 20, (ws.max_row or 0) + 1)):
            v = _clean(ws.cell(row=row, column=col).value).strip().lower()
            if v:
                vals.add(v)
        if vals and vals.issubset({"male", "female", "m", "f", "boy", "girl", ""}):
            return col
    return None


def parse_roster_sheet(ws: Worksheet, branch_name: str, branch_id: int) -> tuple[list[Athlete], list[str]]:
    errors: list[str] = []
    athletes: list[Athlete] = []

    header_row, col_map = _find_header_row(ws)
    if header_row is None or "id" not in col_map:
        errors.append(f"Could not find valid header row in '{ws.title}'")
        return athletes, errors

    # Detect gender column if missing from headers
    gender_col = _detect_gender_column(ws, header_row, col_map)
    if gender_col and "gender" not in col_map:
        col_map["gender"] = gender_col

    # Resolve sessions ambiguity: some branches use "sessions" for season name
    # If sessions column values look like month names, it's actually season
    if "sessions" in col_map:
        sample_vals = set()
        for row in range(header_row + 1, min(header_row + 10, (ws.max_row or 0) + 1)):
            v = _clean(ws.cell(row=row, column=col_map["sessions"]).value).lower()
            if v:
                sample_vals.add(v)
        months = {"jan", "feb", "mar", "apr", "may", "june", "july", "aug", "sep", "oct", "nov", "dec"}
        if sample_vals and sample_vals.issubset(months | {""}):
            # This is a season column, not sessions
            if "season" not in col_map:
                col_map["season"] = col_map.pop("sessions")
            else:
                del col_map["sessions"]

    for row_idx in range(header_row + 1, (ws.max_row or 0) + 1):
        def cell(field: str) -> str:
            col = col_map.get(field)
            if col is None:
                return ""
            return _clean(ws.cell(row=row_idx, column=col).value)

        id_val = cell("id")
        athlete_number = _parse_athlete_number(id_val)
        if athlete_number is None:
            continue

        name = cell("name")
        if not name:
            continue

        # DOB
        dob_raw = ws.cell(row=row_idx, column=col_map["dob"]).value if "dob" in col_map else None
        dob_str = _parse_dob(dob_raw)

        # Pay
        pay_val = _clean_pay(cell("pay"))

        # Gender
        gender = _normalize_gender(cell("gender"))

        # Segment: normalize Member/Non Member → Outsider convention
        segment_raw = cell("segment") or cell("knowing_method")
        if segment_raw:
            seg_lower = segment_raw.strip().lower()
            if seg_lower in ("member", "gems sch.", "gems"):
                segment_raw = "Gems Sch."
            elif seg_lower in ("non member", "outsider"):
                segment_raw = "Outsider"

        athlete = Athlete(
            branch=branch_name,
            branch_id=branch_id,
            athlete_number=athlete_number,
            name=name,
            date_of_birth=dob_str,
            age_computed=_compute_age(dob_str),
            gender=gender,
            step=cell("step") or None,
            type=cell("type") or None,
            days=cell("days") or None,
            sessions=cell("sessions") or None,
            pay=pay_val,
            phone1=cell("phone1") or None,
            phone2=cell("phone2") or None,
            segment=segment_raw or None,
            comment=cell("comment") or None,
            receipt_no=cell("receipt_no") or None,
        )
        athletes.append(athlete)

    return athletes, errors


@dataclass
class AttendanceExtra:
    pay: Optional[str] = None
    receipt_no: Optional[str] = None
    comment: Optional[str] = None


def parse_attendance_sheet(ws: Worksheet, day_pair: str) -> tuple[dict[int, list[ScheduleEntry]], dict[int, AttendanceExtra], list[str]]:
    """Parse attendance sheet — handles all layout variations."""
    errors: list[str] = []
    schedule_map: dict[int, list[ScheduleEntry]] = {}
    extras_map: dict[int, AttendanceExtra] = {}

    max_row = ws.max_row or 0
    max_col = ws.max_column or 0
    if max_row < 2 or max_col < 4:
        return schedule_map, extras_map, errors

    # Detect column layout from first header-like row
    coach_col: Optional[int] = None
    id_col: Optional[int] = None
    name_col: Optional[int] = None
    pay_col: Optional[int] = None
    type_col: Optional[int] = None
    step_col: Optional[int] = None
    comment_col: Optional[int] = None
    pr_col: Optional[int] = None

    for scan_row in range(1, min(6, max_row + 1)):
        for col in range(1, min(40, max_col + 1)):
            val = _clean(ws.cell(row=scan_row, column=col).value).lower().strip()
            if val == "coach" and coach_col is None:
                coach_col = col
            elif val == "id" and id_col is None:
                id_col = col
            elif val == "name" and name_col is None:
                name_col = col
            elif val in ("pay", "fees") and pay_col is None:
                pay_col = col
            elif val == "type" and type_col is None:
                type_col = col
            elif val in ("st.", "step") and step_col is None:
                step_col = col
            elif val in ("comment", "comments") and comment_col is None:
                comment_col = col
            elif val in ("pr.", "pr", "receipt", "receipt num", "receipt no.") and pr_col is None:
                pr_col = col

    if id_col is None and name_col is None:
        return schedule_map, extras_map, errors

    current_time_block: Optional[str] = None
    current_coach: Optional[str] = None

    for row_idx in range(1, max_row + 1):
        # Read key columns
        coach_val = _clean(ws.cell(row=row_idx, column=coach_col).value) if coach_col else ""
        id_val = _clean(ws.cell(row=row_idx, column=id_col).value) if id_col else ""
        name_val = _clean(ws.cell(row=row_idx, column=name_col).value) if name_col else ""

        # Skip empty rows
        if not coach_val and not id_val and not name_val:
            continue

        # Skip repeated header rows (Coach/ID/Name/Pay in the cells)
        if coach_val.lower() in ("coach", "coach ") and id_val.lower() in ("id", ""):
            # But read the time block from the row above (row - 1) or two above
            continue

        # Detect time block rows (e.g., "5:15 to 6:10 pm", "4.15 to 5.10 pm")
        for check_val in [coach_val, id_val, name_val]:
            time_match = re.search(r"\d{1,2}[.:]\d{2}\s*(to|–|-|To)\s*\d{1,2}[.:]\d{2}", check_val, re.IGNORECASE)
            if time_match:
                current_time_block = time_match.group(0).strip()
                break
        else:
            # Not a time block — try to parse as athlete row
            athlete_num = _parse_athlete_number(id_val)
            if athlete_num is None:
                # Maybe this is a section label (team name, etc.)
                continue

            # ID looks valid — this is an athlete row
            # Get coach from coach column on this row
            if coach_col and coach_val and coach_val.lower() not in ("coach", ""):
                current_coach = coach_val

            entry = ScheduleEntry(
                coach=current_coach,
                time_block=current_time_block,
                day_pair=day_pair,
            )
            schedule_map.setdefault(athlete_num, []).append(entry)

            # Extract pay/receipt/comment extras
            pay_raw = None
            if pay_col:
                pay_raw = _clean_pay(_clean(ws.cell(row=row_idx, column=pay_col).value))
            receipt_raw = None
            if pr_col:
                rv = _clean(ws.cell(row=row_idx, column=pr_col).value)
                if rv and rv.lower() not in ("pr.", "pr", "receipt", "0", "0.0"):
                    receipt_raw = rv
            comment_raw = None
            if comment_col:
                cv = _clean(ws.cell(row=row_idx, column=comment_col).value)
                if cv and cv.lower() not in ("comment", "comments"):
                    comment_raw = cv

            if pay_raw or receipt_raw or comment_raw:
                existing = extras_map.get(athlete_num, AttendanceExtra())
                if pay_raw:
                    existing.pay = pay_raw
                if receipt_raw:
                    existing.receipt_no = receipt_raw
                if comment_raw:
                    existing.comment = comment_raw
                extras_map[athlete_num] = existing

    return schedule_map, extras_map, errors


def parse_skills_sheet(ws: Worksheet) -> tuple[dict, list[str], list[str]]:
    errors: list[str] = []
    price_matrix: dict = {}
    coaches: list[str] = []

    max_row = ws.max_row or 0
    max_col = ws.max_column or 0

    for row_idx in range(1, min(15, max_row + 1)):
        row_label = _clean(ws.cell(row=row_idx, column=1).value)
        if row_label and any(kw in row_label.lower() for kw in ["class", "private", "semi", "team", "elite"]):
            prices = {}
            for col_idx in range(2, min(12, max_col + 1)):
                val = _clean(ws.cell(row=row_idx, column=col_idx).value)
                if val and re.match(r"^\d+", val):
                    header = _clean(ws.cell(row=1, column=col_idx).value)
                    if header:
                        prices[header] = val
            if prices:
                price_matrix[row_label] = prices

    for col_idx in range(18, min(25, max_col + 1)):
        for row_idx in range(1, max_row + 1):
            val = _clean(ws.cell(row=row_idx, column=col_idx).value)
            if val and re.search(r"[a-zA-Z]", val) and len(val) > 2:
                if val.lower() not in ("coach", "coaches", "name", "total"):
                    coaches.append(val)

    return price_matrix, coaches, errors


def _is_attendance_sheet(name: str) -> bool:
    """Check if a sheet name looks like an attendance/schedule sheet."""
    n = name.strip().lower()
    # Skip roster and skills sheets
    if "reg" in n or n.startswith("sk"):
        return False
    # Match known attendance patterns
    attendance_keywords = ["sat", "sun", "mon", "tue", "wed", "thu", "fri",
                          "team", "baby", "pre", "dax", "k.baby"]
    return any(kw in n for kw in attendance_keywords)


def parse_workbook(file_path: str, branch_name: str, branch_id: int) -> tuple[list[Athlete], list[str], dict, list[str]]:
    """Parse a complete branch workbook. Returns (athletes, coaches, price_matrix, errors)."""
    errors: list[str] = []
    athletes: list[Athlete] = []
    coaches: list[str] = []
    price_matrix: dict = {}

    try:
        wb = load_workbook(file_path, read_only=False, data_only=True)
    except Exception as e:
        errors.append(f"Failed to open workbook: {e}")
        return athletes, coaches, price_matrix, errors

    roster_sheet = None
    attendance_sheets: list[tuple[Worksheet, str]] = []
    skills_sheet = None

    for sheet_name in wb.sheetnames:
        name = sheet_name.strip().lower()
        if "reg" in name:
            roster_sheet = wb[sheet_name]
        elif name.startswith("sk"):
            skills_sheet = wb[sheet_name]
        elif _is_attendance_sheet(sheet_name):
            attendance_sheets.append((wb[sheet_name], sheet_name.strip()))

    if roster_sheet is None:
        errors.append("No roster sheet found (expected sheet with 'Reg' in name)")
        wb.close()
        return athletes, coaches, price_matrix, errors

    # Parse roster
    athletes, roster_errors = parse_roster_sheet(roster_sheet, branch_name, branch_id)
    errors.extend(roster_errors)
    logger.info(f"Roster '{roster_sheet.title}': {len(athletes)} athletes")

    # Build athlete lookup
    athlete_by_num: dict[int, Athlete] = {a.athlete_number: a for a in athletes}

    # Parse attendance sheets — merge schedule + extras
    for ws, day_pair in attendance_sheets:
        try:
            schedule_map, extras_map, att_errors = parse_attendance_sheet(ws, day_pair)
            errors.extend(att_errors)

            for num, entries in schedule_map.items():
                if num in athlete_by_num:
                    athlete_by_num[num].schedule.extend(entries)

            for num, extra in extras_map.items():
                if num in athlete_by_num:
                    a = athlete_by_num[num]
                    if extra.pay and not a.pay:
                        a.pay = extra.pay
                    if extra.receipt_no and not a.receipt_no:
                        a.receipt_no = extra.receipt_no
                    if extra.comment and not a.comment:
                        a.comment = extra.comment

            sched_count = sum(len(v) for v in schedule_map.values())
            if sched_count > 0:
                logger.info(f"  Attendance '{day_pair}': {sched_count} schedule entries, {len(extras_map)} extras")
        except Exception as e:
            errors.append(f"Error parsing attendance sheet '{day_pair}': {e}")

    # Parse Sk. sheet
    if skills_sheet:
        try:
            price_matrix, coaches, sk_errors = parse_skills_sheet(skills_sheet)
            errors.extend(sk_errors)
        except Exception as e:
            errors.append(f"Error parsing Sk. sheet: {e}")

    wb.close()
    return athletes, coaches, price_matrix, errors
