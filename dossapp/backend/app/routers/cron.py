"""Cron endpoints — invoked by Vercel Cron (Authorization: Bearer $CRON_SECRET).

An admin JWT is also accepted so triggers can be run manually from the app/API.
"""

import hmac
import logging
from datetime import datetime

from fastapi import APIRouter, Header, HTTPException
from typing import Optional

from app.config import settings
from app.database import async_session
from app.utils.auth import decode_token

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/cron", tags=["cron"])


def _authorize(authorization: Optional[str]):
    token = (authorization or "").removeprefix("Bearer ").strip()
    if not token:
        raise HTTPException(status_code=401, detail="Missing authorization")

    if settings.cron_secret and hmac.compare_digest(token, settings.cron_secret):
        return

    payload = decode_token(token)
    if payload and payload.get("type") == "access" and payload.get("role") == "admin":
        return

    raise HTTPException(status_code=401, detail="Invalid cron credentials")


def _get_roster_source():
    from app.main import roster_source
    return roster_source


@router.get("/notifications")
async def cron_notifications(force: bool = False, authorization: Optional[str] = Header(None)):
    """Refresh rosters and run schedule-change + missed-session triggers."""
    _authorize(authorization)
    from app.services.push_triggers import run_roster_triggers

    source = _get_roster_source()
    await source.refresh_all(force=force)
    rosters = await source.get_all_rosters()

    totals = {"schedule": 0, "missed": 0}
    for branch_id, roster in rosters.items():
        async with async_session() as db:
            counts = await run_roster_triggers(roster, db)
        totals["schedule"] += counts["schedule"]
        totals["missed"] += counts["missed"]

    logger.info(f"Cron notifications: {totals}")
    return {"branches": len(rosters), "sent": totals}


@router.get("/payment-reminders")
async def cron_payment_reminders(
    force: bool = False,
    authorization: Optional[str] = Header(None),
):
    """Send unpaid reminders — only on the configured day of month (default 25th)."""
    _authorize(authorization)

    if not force and datetime.now().day != settings.payment_reminder_day:
        return {"skipped": True, "reason": f"Not day {settings.payment_reminder_day}"}

    from app.services.push_triggers import send_payment_reminders

    source = _get_roster_source()
    rosters = await source.get_all_rosters()

    async with async_session() as db:
        sent = await send_payment_reminders(db, rosters)

    logger.info(f"Cron payment reminders: sent={sent}")
    return {"branches": len(rosters), "sent": sent}


@router.get("/debug-excel-structure")
async def debug_excel_structure(
    branch_id: int = 1,
    authorization: Optional[str] = Header(None),
):
    """Temporary debug endpoint: dump raw Excel structure for a branch."""
    _authorize(authorization)

    import asyncio
    import os

    # Get the branch config from the roster source
    source = _get_roster_source()
    config = source._configs.get(branch_id)
    if not config:
        raise HTTPException(status_code=404, detail=f"No config for branch {branch_id}")

    # Download the file
    local_path = config.get("local_file_path")
    if local_path:
        tmp_path = local_path
        cleanup = False
    else:
        drive_file_id = config.get("drive_file_id")
        if not drive_file_id:
            raise HTTPException(status_code=400, detail="No drive_file_id configured")
        from app.services.drive_reader import download_file
        tmp_path = await asyncio.to_thread(download_file, drive_file_id)
        if not tmp_path:
            raise HTTPException(status_code=500, detail="Failed to download Excel file")
        cleanup = True

    try:
        from openpyxl import load_workbook

        wb = load_workbook(tmp_path, data_only=True)
        result = {"branch_id": branch_id, "branch_name": config.get("branch_name"), "sheets": []}

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            max_row = ws.max_row or 0
            max_col = ws.max_column or 0

            # Determine how many rows to dump
            dump_rows = min(20, max_row) if max_row else 0

            # For the roster sheet ("Reg"), also flag it
            is_roster = "reg" in sheet_name.lower()

            cells = []
            for row_idx in range(1, dump_rows + 1):
                for col_idx in range(1, max_col + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    val = cell.value
                    cells.append({
                        "row": row_idx,
                        "col": col_idx,
                        "value": str(val) if val is not None else None,
                        "type": type(val).__name__ if val is not None else "NoneType",
                    })

            sheet_info = {
                "sheet_name": sheet_name,
                "max_row": max_row,
                "max_column": max_col,
                "is_roster_sheet": is_roster,
                "dumped_rows": dump_rows,
                "cells": cells,
            }

            # For roster sheet, dump the full header row even if > 20 cols
            if is_roster and max_row >= 1:
                header_row = []
                for col_idx in range(1, max_col + 1):
                    cell = ws.cell(row=1, column=col_idx)
                    val = cell.value
                    header_row.append({
                        "col": col_idx,
                        "value": str(val) if val is not None else None,
                    })
                sheet_info["full_header_row"] = header_row

            result["sheets"].append(sheet_info)

        wb.close()
        return result

    finally:
        if cleanup:
            try:
                os.unlink(tmp_path)
            except OSError:
                pass
