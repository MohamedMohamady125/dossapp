"""Parse Aqua Athletic Excel workbooks into Athlete objects.

Universal parser — handles all branch format variations:
- Rehab/Madinaty: 19-col layout with M Code + visa, M-prefix IDs, M/F gender
- Choueifat/Choueifat Cairo: 17-col layout, C/E-prefix IDs, Male/Female gender
- Different header names: pay/Pay/Fees/fees, Date of Birth/Birth Date/B . D, etc.
- Attendance sheets with coach in Col1, repeated sub-headers, Pay in attendance
"""

import logging
import re
from dataclasses import dataclass
from datetime import datetime, timedelta
from typing import Optional

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.services.roster_source import Athlete, ScheduleEntry

logger = logging.getLogger(__name__)

EXCEL_EPOCH = datetime(1899, 12, 30)

# Map of lowercase/stripped header → field name (all known variations)
HEADER_MAP = {
    "f": "segment", "membership": "segment", "member": "segment",
    "id": "id",
    "name": "name",
    "date of birth": "dob", "birth date": "dob", "dob": "dob", "b . d": "dob", "b.d": "dob",
    "gender": "gender",
    "step": "step", "level": "step", "st.": "step",
    "type": "type",
    "days": "days",
    "times": "times",
    "season": "season",
    "sessions": "sessions",
    "pay": "pay", "fees": "pay", "amount": "pay",
    "phone 1": "phone1", "phone": "phone1", "phone1": "phone1",
    "phone 2": "phone2", "phone2": "phone2",
    "receipt no.": "receipt_no", "receipt num": "receipt_no",
    "receipt number": "receipt_no", "receipt no": "receipt_no",
    "comment": "comment", "comments": "comment", "notes": "comment",
    "visa": "visa",
    "knowing method": "knowing_method",
    "m code": "m_code",
}


def _clean(val) -> str:
    if val is None:
        return ""
    s = str(val).strip()
    if s in ("#REF!", "#N/A", "#VALUE!", "#DIV/0!"):
        return ""
    return s


def _parse_athlete_number(id_val: str) -> Optional[int]:
    """Extract number from ID like 'M31', 'C1001', 'E1002' → number."""
    if not id_val:
        return None
    nums = re.findall(r"\d+", id_val)
    if nums:
        return int(nums[0])
    return None


def _parse_dob(raw) -> Optional[str]:
    """Convert various DOB formats to ISO date string."""
    if raw is None:
        return None
    if isinstance(raw, datetime):
        return raw.strftime("%Y-%m-%d")
    s = str(raw).strip()
    if not s:
        return None
    # Try Excel serial number
    try:
        num = float(s)
        if 1 < num < 100000:
            dt = EXCEL_EPOCH + timedelta(days=num)
            return dt.strftime("%Y-%m-%d")
    except (ValueError, TypeError, OverflowError):
        pass
    # Try various string date formats
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d/%m/%y", "%d\\%m\\%Y",
                "%d\\%m\\%y", "%m/%d/%Y", "%m/%d/%y"):
        try:
            dt = datetime.strptime(s, fmt)
            if dt.year > 2050:
                dt = dt.replace(year=dt.year - 100)
            return dt.strftime("%Y-%m-%d")
        except ValueError:
            pass
    # Handle weird separators like 'L', '.'
    cleaned = re.sub(r"[LlBb.\\/-]", "/", s)
    for fmt in ("%d/%m/%Y", "%d/%m/%y"):
        try:
            dt = datetime.strptime(cleaned, fmt)
            if dt.year > 2050:
                dt = dt.replace(year=dt.year - 100)
            return dt.strftime("%Y-%m-%d")
        except ValueError:
            pass
    return None


def _compute_age(dob_str: Optional[str]) -> Optional[float]:
    if not dob_str:
        return None
    try:
        dob = datetime.strptime(dob_str, "%Y-%m-%d")
        age = (datetime.now() - dob).days / 365.25
        if age < 0 or age > 100:
            return None
        return round(age, 1)
    except (ValueError, TypeError):
        return None


def _normalize_gender(raw: str) -> Optional[str]:
    if not raw:
        return None
    g = raw.strip().lower()
    if g in ("m", "male", "boy", "ذكر"):
        return "M"
    if g in ("f", "female", "girl", "أنثى"):
        return "F"
    return raw.strip() or None


def _clean_phone(raw) -> Optional[str]:
    """Clean phone value from Excel — handles float numbers, leading zero restoration."""
    if raw is None:
        return None
    # If openpyxl gives us a float (e.g. 1273604569.0), convert to int first
    if isinstance(raw, float):
        raw = str(int(raw))
    s = str(raw).strip()
    if not s:
        return None
    # Strip trailing .0 from string representation
    if s.endswith(".0"):
        s = s[:-2]
    # Remove non-digit characters except leading +
    if s.startswith("+"):
        digits = "+" + re.sub(r"[^\d]", "", s[1:])
    else:
        digits = re.sub(r"[^\d]", "", s)
    if not digits or len(digits) < 7:
        return None
    # Egyptian numbers: if 10 digits starting with 1 (missing leading 0), add it
    if len(digits) == 10 and digits.startswith("1"):
        digits = "0" + digits
    return digits


def _clean_pay(raw: str) -> Optional[str]:
    """Clean pay value — return numeric string or None."""
    if not raw:
        return None
    s = raw.strip().lower()
    # Skip non-numeric values like 'Trans', 'Pay', header repeats
    if s in ("pay", "fees", "trans", "trans ", "0", "0.0", ""):
        return None
    cleaned = re.sub(r"[^\d.]", "", s)
    if not cleaned:
        return None
    try:
        float(cleaned)
        return cleaned
    except ValueError:
        return None


def _find_header_row(ws: Worksheet) -> tuple[Optional[int], dict[str, int]]:
    """Find the header row and map field names to column indices."""
    max_row = min(10, ws.max_row or 10)
    max_col = min(50, ws.max_column or 50)

    for row_idx in range(1, max_row + 1):
        cells = {}
        for col_idx in range(1, max_col + 1):
            val = _clean(ws.cell(row=row_idx, column=col_idx).value).lower().strip()
            if val:
                cells[val] = col_idx

        # Match known headers
        matches = {}
        for header_text, field_name in HEADER_MAP.items():
            if header_text in cells and field_name not in matches:
                matches[field_name] = cells[header_text]

        # Need at least ID + Name to consider this a valid header row
        if "id" in matches and "name" in matches:
            return row_idx, matches

    # Fallback: try to detect by column position if no clear headers
    return None, {}


def _detect_gender_column(ws: Worksheet, header_row: int, col_map: dict) -> Optional[int]:
    """Some branches (Choueifat Cairo) have no gender header — detect by values."""
    if "gender" in col_map:
        return col_map["gender"]
    # Check columns near where gender typically is (5-8)
    for col in range(5, 9):
        if col in col_map.values():
            continue
        # Check first 5 data rows
        vals = set()
        for row in range(header_row + 1, min(header_row + 20, (ws.max_row or 0) + 1)):
            v = _clean(ws.cell(row=row, column=col).value).strip().lower()
            if v:
                vals.add(v)
        if vals and vals.issubset({"male", "female", "m", "f", "boy", "girl", ""}):
            return col
    return None


def parse_roster_sheet(ws: Worksheet, branch_name: str, branch_id: int) -> tuple[list[Athlete], list[str]]:
    errors: list[str] = []
    athletes: list[Athlete] = []

    header_row, col_map = _find_header_row(ws)
    if header_row is None or "id" not in col_map:
        errors.append(f"Could not find valid header row in '{ws.title}'")
        return athletes, errors

    # Detect gender column if missing from headers
    gender_col = _detect_gender_column(ws, header_row, col_map)
    if gender_col and "gender" not in col_map:
        col_map["gender"] = gender_col

    # Resolve sessions ambiguity: some branches use "sessions" for season name
    # If sessions column values look like month names, it's actually season
    if "sessions" in col_map:
        sample_vals = set()
        for row in range(header_row + 1, min(header_row + 10, (ws.max_row or 0) + 1)):
            v = _clean(ws.cell(row=row, column=col_map["sessions"]).value).lower()
            if v:
                sample_vals.add(v)
        months = {"jan", "feb", "mar", "apr", "may", "june", "july", "aug", "sep", "oct", "nov", "dec"}
        if sample_vals and sample_vals.issubset(months | {""}):
            # This is a season column, not sessions
            if "season" not in col_map:
                col_map["season"] = col_map.pop("sessions")
            else:
                del col_map["sessions"]

    consecutive_empty = 0
    for row_idx in range(header_row + 1, (ws.max_row or 0) + 1):
        def cell(field: str) -> str:
            col = col_map.get(field)
            if col is None:
                return ""
            return _clean(ws.cell(row=row_idx, column=col).value)

        id_val = cell("id")
        athlete_number = _parse_athlete_number(id_val)
        if athlete_number is None:
            consecutive_empty += 1
            # Stop after 50 consecutive empty rows — avoids scanning 50k empty rows
            if consecutive_empty > 50:
                break
            continue

        consecutive_empty = 0
        name = cell("name")
        if not name:
            continue

        # DOB
        dob_raw = ws.cell(row=row_idx, column=col_map["dob"]).value if "dob" in col_map else None
        dob_str = _parse_dob(dob_raw)

        # Pay
        pay_val = _clean_pay(cell("pay"))

        # Gender
        gender = _normalize_gender(cell("gender"))

        # Segment: normalize Member/Non Member → Outsider convention
        segment_raw = cell("segment") or cell("knowing_method")
        if segment_raw:
            seg_lower = segment_raw.strip().lower()
            if seg_lower in ("member", "gems sch.", "gems"):
                segment_raw = "Gems Sch."
            elif seg_lower in ("non member", "outsider"):
                segment_raw = "Outsider"

        # Phone: read raw cell value to handle float numbers properly
        phone1_raw = ws.cell(row=row_idx, column=col_map["phone1"]).value if "phone1" in col_map else None
        phone2_raw = ws.cell(row=row_idx, column=col_map["phone2"]).value if "phone2" in col_map else None

        athlete = Athlete(
            branch=branch_name,
            branch_id=branch_id,
            athlete_number=athlete_number,
            name=name,
            date_of_birth=dob_str,
            age_computed=_compute_age(dob_str),
            gender=gender,
            step=cell("step") or None,
            type=cell("type") or None,
            days=cell("days") or None,
            sessions=cell("sessions") or None,
            pay=pay_val,
            phone1=_clean_phone(phone1_raw),
            phone2=_clean_phone(phone2_raw),
            segment=segment_raw or None,
            comment=cell("comment") or None,
            receipt_no=cell("receipt_no") or None,
        )
        athletes.append(athlete)

    return athletes, errors


@dataclass
class AttendanceExtra:
    pay: Optional[str] = None
    receipt_no: Optional[str] = None
    comment: Optional[str] = None
    type: Optional[str] = None
    step: Optional[str] = None
    sessions: Optional[str] = None


# Attendance mark classification (lowercased). Any other non-empty mark
# defaults to Present — conservative, avoids false "missed sessions" alerts.
ABSENT_MARKS = {"a", "ab", "abs", "absent", "x", "✗", "✘", "غ", "غياب", "0"}
PRESENT_MARKS = {"p", "present", "✓", "✔", "1", "h", "attended"}


def _classify_attendance_mark(raw) -> Optional[str]:
    """Classify a cell mark as 'P'/'A', or None for empty (no session recorded)."""
    s = _clean(raw).lower()
    if not s:
        return None
    if s in ABSENT_MARKS:
        return "A"
    return "P"


def _header_cell_to_date(val, default_year: int, default_month: Optional[int] = None) -> Optional[str]:
    """Interpret a header cell as a session date. Returns ISO date or None.

    Handles: datetime objects, Excel serial numbers, d/m/y strings, d/m strings,
    and bare day-of-month numbers (1-31) when default_month is provided.
    """
    if val is None:
        return None
    if isinstance(val, datetime):
        if 2000 <= val.year <= 2100:
            return val.strftime("%Y-%m-%d")
        return None
    s = str(val).strip()
    if not s:
        return None
    # Try as a number first
    try:
        num = float(s)
        # Excel serial number (roughly 2005–2060)
        if 38000 < num < 60000:
            return (EXCEL_EPOCH + timedelta(days=num)).strftime("%Y-%m-%d")
        # Day-of-month number (1–31) — needs a default_month context
        if default_month and 1 <= num <= 31 and num == int(num):
            try:
                return datetime(default_year, default_month, int(num)).strftime("%Y-%m-%d")
            except ValueError:
                return None
        return None
    except (ValueError, TypeError, OverflowError):
        pass
    # String date with year: 5/10/2026, 5-10-26, 5.10.2026
    m = re.fullmatch(r"(\d{1,2})[/\-.](\d{1,2})[/\-.](\d{2,4})", s)
    if m:
        d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if y < 100:
            y += 2000
        try:
            return datetime(y, mo, d).strftime("%Y-%m-%d")
        except ValueError:
            return None
    # Day/month only: 5/10, 5-10 — infer year
    m = re.fullmatch(r"(\d{1,2})[/\-.](\d{1,2})", s)
    if m:
        d, mo = int(m.group(1)), int(m.group(2))
        try:
            return datetime(default_year, mo, d).strftime("%Y-%m-%d")
        except ValueError:
            return None
    return None


def _detect_date_columns(ws: Worksheet, max_col: int, known_cols: set[int], default_month: Optional[int] = None) -> dict[int, str]:
    """Find per-session date columns in the header area. Returns {col: ISO date}."""
    default_year = datetime.now().year
    date_cols: dict[int, str] = {}
    max_row = ws.max_row or 0
    for scan_row in range(1, min(6, max_row + 1)):
        for col in range(1, min(60, max_col + 1)):
            if col in known_cols or col in date_cols:
                continue
            iso = _header_cell_to_date(ws.cell(row=scan_row, column=col).value, default_year, default_month)
            if iso:
                date_cols[col] = iso
    return date_cols


def parse_attendance_sheet(ws: Worksheet, day_pair: str, default_month: Optional[int] = None) -> tuple[dict[int, list[ScheduleEntry]], dict[int, AttendanceExtra], dict[int, dict[str, str]], list[str]]:
    """Parse attendance sheet — handles all layout variations.

    Returns (schedule_map, extras_map, attendance_map, errors) where
    attendance_map is {athlete_number: {iso_date: 'P'|'A'}}.
    """
    errors: list[str] = []
    schedule_map: dict[int, list[ScheduleEntry]] = {}
    extras_map: dict[int, AttendanceExtra] = {}
    attendance_map: dict[int, dict[str, str]] = {}

    max_row = ws.max_row or 0
    max_col = ws.max_column or 0
    if max_row < 2 or max_col < 4:
        return schedule_map, extras_map, attendance_map, errors

    # Detect column layout from first header-like row
    all_coach_cols: list[int] = []
    id_col: Optional[int] = None
    name_col: Optional[int] = None
    pay_col: Optional[int] = None
    type_col: Optional[int] = None
    step_col: Optional[int] = None
    comment_col: Optional[int] = None
    pr_col: Optional[int] = None

    for scan_row in range(1, min(6, max_row + 1)):
        for col in range(1, min(40, max_col + 1)):
            val = _clean(ws.cell(row=scan_row, column=col).value).lower().strip()
            if val == "coach":
                all_coach_cols.append(col)
            elif val == "id" and id_col is None:
                id_col = col
            elif val == "name" and name_col is None:
                name_col = col
            elif val in ("pay", "fees") and pay_col is None:
                pay_col = col
            elif val == "type" and type_col is None:
                type_col = col
            elif val in ("st.", "st", "step") and step_col is None:
                step_col = col
            elif val in ("comment", "comments") and comment_col is None:
                comment_col = col
            elif val in ("pr.", "pr", "receipt", "receipt num", "receipt no.") and pr_col is None:
                pr_col = col

    # Use the coach column closest to (just before) the Type or ID column —
    # attendance sheets often have multiple "Coach" headers (spare coach, attendance, actual coach)
    coach_col: Optional[int] = None
    if all_coach_cols:
        ref_col = type_col or id_col or 999
        # Prefer the coach column immediately before Type/ID
        before = [c for c in all_coach_cols if c < ref_col]
        coach_col = max(before) if before else all_coach_cols[-1]

    logger.info(f"  Attendance '{day_pair}': layout id_col={id_col} name_col={name_col} coach_col={coach_col} type_col={type_col} step_col={step_col} pay_col={pay_col} max_row={max_row} max_col={max_col}")

    if id_col is None and name_col is None:
        logger.warning(f"  Attendance '{day_pair}': NO id or name column found — skipping")
        # Dump first 3 rows to help debug
        for r in range(1, min(4, max_row + 1)):
            cells = {col: repr(ws.cell(row=r, column=col).value) for col in range(1, min(20, max_col + 1)) if ws.cell(row=r, column=col).value is not None}
            logger.warning(f"    Row {r}: {cells}")
        return schedule_map, extras_map, attendance_map, errors

    # Detect per-session date columns (attendance marks live under them)
    known_cols = {c for c in (coach_col, id_col, name_col, pay_col, type_col, step_col, comment_col, pr_col) if c}
    date_cols = _detect_date_columns(ws, max_col, known_cols, default_month=default_month)

    # Log date detection details — dump header cells beyond known columns for debugging
    if not date_cols:
        logger.warning(f"  Attendance '{day_pair}': NO date columns detected! Dumping header cells:")
        for r in range(1, min(6, max_row + 1)):
            cells = {}
            for col in range(1, min(60, max_col + 1)):
                v = ws.cell(row=r, column=col).value
                if v is not None and col not in known_cols:
                    cells[col] = f"{type(v).__name__}:{repr(v)}"
            if cells:
                logger.warning(f"    Row {r} non-header cells: {cells}")
    else:
        logger.info(f"  Attendance '{day_pair}': detected {len(date_cols)} date columns: {sorted(date_cols.values())}")

    current_time_block: Optional[str] = None
    current_coach: Optional[str] = None

    for row_idx in range(1, max_row + 1):
        # Read key columns
        coach_val = _clean(ws.cell(row=row_idx, column=coach_col).value) if coach_col else ""
        id_val = _clean(ws.cell(row=row_idx, column=id_col).value) if id_col else ""
        name_val = _clean(ws.cell(row=row_idx, column=name_col).value) if name_col else ""

        # Skip empty rows
        if not coach_val and not id_val and not name_val:
            continue

        # Skip repeated header rows (Coach/ID/Name/Pay in the cells)
        if coach_val.lower() in ("coach", "coach ") and id_val.lower() in ("id", ""):
            # But read the time block from the row above (row - 1) or two above
            continue

        # Detect time block rows (e.g., "5:15 to 6:10 pm", "4.15 to 5.10 pm")
        for check_val in [coach_val, id_val, name_val]:
            time_match = re.search(r"\d{1,2}[.:]\d{2}\s*(to|–|-|To)\s*\d{1,2}[.:]\d{2}", check_val, re.IGNORECASE)
            if time_match:
                current_time_block = time_match.group(0).strip()
                break
        else:
            # Not a time block — try to parse as athlete row
            athlete_num = _parse_athlete_number(id_val)
            if athlete_num is None:
                # Maybe this is a section label (team name, etc.)
                continue

            # ID looks valid — this is an athlete row
            # Get coach from coach column on this row
            if coach_col and coach_val and coach_val.lower() not in ("coach", ""):
                current_coach = coach_val

            entry = ScheduleEntry(
                coach=current_coach,
                time_block=current_time_block,
                day_pair=day_pair,
            )
            schedule_map.setdefault(athlete_num, []).append(entry)

            # Attendance marks under date columns
            for col, iso_date in date_cols.items():
                mark = _classify_attendance_mark(ws.cell(row=row_idx, column=col).value)
                if mark:
                    attendance_map.setdefault(athlete_num, {})[iso_date] = mark

            # Extract type/step from attendance row
            type_raw = None
            if type_col:
                tv = _clean(ws.cell(row=row_idx, column=type_col).value)
                if tv and tv.lower() not in ("type", ""):
                    type_raw = tv
            step_raw = None
            if step_col:
                sv = _clean(ws.cell(row=row_idx, column=step_col).value)
                if sv and sv.lower() not in ("st.", "st", "step", "level", ""):
                    step_raw = sv

            # Extract pay/receipt/comment extras
            pay_raw = None
            if pay_col:
                pay_raw = _clean_pay(_clean(ws.cell(row=row_idx, column=pay_col).value))
            receipt_raw = None
            if pr_col:
                rv = _clean(ws.cell(row=row_idx, column=pr_col).value)
                if rv and rv.lower() not in ("pr.", "pr", "receipt", "0", "0.0"):
                    receipt_raw = rv
            comment_raw = None
            if comment_col:
                cv = _clean(ws.cell(row=row_idx, column=comment_col).value)
                if cv and cv.lower() not in ("comment", "comments"):
                    comment_raw = cv

            if pay_raw or receipt_raw or comment_raw or type_raw or step_raw:
                existing = extras_map.get(athlete_num, AttendanceExtra())
                if pay_raw:
                    existing.pay = pay_raw
                if receipt_raw:
                    existing.receipt_no = receipt_raw
                if comment_raw:
                    existing.comment = comment_raw
                if type_raw and not existing.type:
                    existing.type = type_raw
                if step_raw and not existing.step:
                    existing.step = step_raw
                extras_map[athlete_num] = existing

    total_marks = sum(len(v) for v in attendance_map.values())
    logger.info(f"  Attendance '{day_pair}' result: {len(schedule_map)} athletes scheduled, {len(attendance_map)} with marks, {total_marks} total marks")

    return schedule_map, extras_map, attendance_map, errors


def parse_skills_sheet(ws: Worksheet) -> tuple[dict, list[str], list[str]]:
    errors: list[str] = []
    price_matrix: dict = {}
    coaches: list[str] = []

    max_row = ws.max_row or 0
    max_col = ws.max_column or 0

    for row_idx in range(1, min(15, max_row + 1)):
        row_label = _clean(ws.cell(row=row_idx, column=1).value)
        if row_label and any(kw in row_label.lower() for kw in ["class", "private", "semi", "team", "elite"]):
            prices = {}
            for col_idx in range(2, min(12, max_col + 1)):
                val = _clean(ws.cell(row=row_idx, column=col_idx).value)
                if val and re.match(r"^\d+", val):
                    header = _clean(ws.cell(row=1, column=col_idx).value)
                    if header:
                        prices[header] = val
            if prices:
                price_matrix[row_label] = prices

    for col_idx in range(18, min(25, max_col + 1)):
        for row_idx in range(1, max_row + 1):
            val = _clean(ws.cell(row=row_idx, column=col_idx).value)
            if val and re.search(r"[a-zA-Z]", val) and len(val) > 2:
                if val.lower() not in ("coach", "coaches", "name", "total"):
                    coaches.append(val)

    return price_matrix, coaches, errors


_MONTH_NAMES = {
    "jan": 1, "january": 1, "feb": 2, "february": 2, "mar": 3, "march": 3,
    "apr": 4, "april": 4, "may": 5, "jun": 6, "june": 6,
    "jul": 7, "july": 7, "aug": 8, "august": 8, "sep": 9, "september": 9,
    "oct": 10, "october": 10, "nov": 11, "november": 11, "dec": 12, "december": 12,
}


def _detect_month_from_name(name: str) -> Optional[int]:
    """Extract month number from a sheet name like 'July. Reg' or 'August Reg'."""
    for word in re.split(r"[\s.\-_]+", name.strip().lower()):
        if word in _MONTH_NAMES:
            return _MONTH_NAMES[word]
    return None


def _is_attendance_sheet(name: str) -> bool:
    """Check if a sheet name looks like an attendance/schedule sheet."""
    n = name.strip().lower()
    # Skip roster and skills sheets
    if "reg" in n or n.startswith("sk"):
        return False
    # Match known attendance patterns
    attendance_keywords = ["sat", "sun", "mon", "tue", "wed", "thu", "fri",
                          "team", "baby", "pre", "dax", "k.baby"]
    return any(kw in n for kw in attendance_keywords)


def parse_workbook(file_path: str, branch_name: str, branch_id: int) -> tuple[list[Athlete], list[str], dict, list[str]]:
    """Parse a complete branch workbook. Returns (athletes, coaches, price_matrix, errors)."""
    errors: list[str] = []
    athletes: list[Athlete] = []
    coaches: list[str] = []
    price_matrix: dict = {}

    try:
        wb = load_workbook(file_path, read_only=False, data_only=True)
    except Exception as e:
        errors.append(f"Failed to open workbook: {e}")
        return athletes, coaches, price_matrix, errors

    roster_sheet = None
    attendance_sheets: list[tuple[Worksheet, str]] = []
    skills_sheet = None

    for sheet_name in wb.sheetnames:
        name = sheet_name.strip().lower()
        if "reg" in name:
            roster_sheet = wb[sheet_name]
        elif name.startswith("sk"):
            skills_sheet = wb[sheet_name]
        elif _is_attendance_sheet(sheet_name):
            attendance_sheets.append((wb[sheet_name], sheet_name.strip()))

    # Detect month from roster sheet name (e.g., "July. Reg" → 7)
    roster_month = _detect_month_from_name(roster_sheet.title if roster_sheet else "")
    logger.info(f"Workbook sheets: {wb.sheetnames}, detected month: {roster_month}")
    logger.info(f"  Roster: {roster_sheet.title if roster_sheet else 'NONE'}, Attendance: {[dp for _, dp in attendance_sheets]}, Skills: {skills_sheet.title if skills_sheet else 'NONE'}")

    if roster_sheet is None:
        errors.append("No roster sheet found (expected sheet with 'Reg' in name)")
        wb.close()
        return athletes, coaches, price_matrix, errors

    # Parse roster
    athletes, roster_errors = parse_roster_sheet(roster_sheet, branch_name, branch_id)
    errors.extend(roster_errors)
    logger.info(f"Roster '{roster_sheet.title}': {len(athletes)} athletes")

    # Build athlete lookup
    athlete_by_num: dict[int, Athlete] = {a.athlete_number: a for a in athletes}

    # Parse attendance sheets — merge schedule + extras
    for ws, day_pair in attendance_sheets:
        try:
            schedule_map, extras_map, attendance_map, att_errors = parse_attendance_sheet(ws, day_pair, default_month=roster_month)
            errors.extend(att_errors)

            for num, entries in schedule_map.items():
                if num in athlete_by_num:
                    athlete_by_num[num].schedule.extend(entries)

            for num, marks in attendance_map.items():
                if num in athlete_by_num:
                    athlete_by_num[num].attendance.update(marks)

            # Count how many type/step values will be filled from attendance
            will_merge_type = sum(1 for n, e in extras_map.items() if n in athlete_by_num and e.type and not athlete_by_num[n].type)
            will_merge_step = sum(1 for n, e in extras_map.items() if n in athlete_by_num and e.step and not athlete_by_num[n].step)

            for num, extra in extras_map.items():
                if num in athlete_by_num:
                    a = athlete_by_num[num]
                    if extra.pay and not a.pay:
                        a.pay = extra.pay
                    if extra.receipt_no and not a.receipt_no:
                        a.receipt_no = extra.receipt_no
                    if extra.comment and not a.comment:
                        a.comment = extra.comment
                    # Fill in type/step from attendance if roster is missing them
                    if extra.type and not a.type:
                        a.type = extra.type
                    if extra.step and not a.step:
                        a.step = extra.step

            sched_count = sum(len(v) for v in schedule_map.values())
            if sched_count > 0 or will_merge_type or will_merge_step:
                logger.info(f"  Attendance '{day_pair}': {sched_count} schedule, {len(extras_map)} extras, will_merge type={will_merge_type} step={will_merge_step}")
        except Exception as e:
            errors.append(f"Error parsing attendance sheet '{day_pair}': {e}")

    # Parse Sk. sheet
    if skills_sheet:
        try:
            price_matrix, coaches, sk_errors = parse_skills_sheet(skills_sheet)
            errors.extend(sk_errors)
        except Exception as e:
            errors.append(f"Error parsing Sk. sheet: {e}")

    wb.close()
    return athletes, coaches, price_matrix, errors
