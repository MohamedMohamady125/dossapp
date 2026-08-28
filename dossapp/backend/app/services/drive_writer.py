"""Write payment data back to Google Drive Excel/Sheets files.

Updates the Pay and Receipt columns in the roster sheet when an athlete is marked as paid.
Uses per-file threading locks to prevent concurrent writes from corrupting data.
"""

import logging
import re
import threading
import time
from typing import Optional

from app.config import settings

logger = logging.getLogger(__name__)

# Per-file locks to serialize writes to the same spreadsheet/Excel file.
# This prevents the download-modify-upload race condition.
_file_locks: dict[str, threading.Lock] = {}
_file_locks_guard = threading.Lock()


def _get_file_lock(file_id: str) -> threading.Lock:
    """Get or create a threading lock for a specific Drive file."""
    with _file_locks_guard:
        if file_id not in _file_locks:
            _file_locks[file_id] = threading.Lock()
        return _file_locks[file_id]


def _retry(func, *args, max_retries: int = 3, **kwargs):
    """Retry a function up to max_retries times with exponential backoff."""
    last_error = None
    for attempt in range(max_retries):
        try:
            return func(*args, **kwargs)
        except Exception as e:
            last_error = e
            if attempt < max_retries - 1:
                wait = (2 ** attempt) * 0.5  # 0.5s, 1s, 2s
                logger.warning(f"Drive write attempt {attempt + 1} failed: {e}, retrying in {wait}s")
                time.sleep(wait)
            else:
                logger.error(f"Drive write failed after {max_retries} attempts: {e}")
    raise last_error

try:
    from google.oauth2 import service_account
    from googleapiclient.discovery import build
    _HAS_GOOGLE = True
except ImportError:
    _HAS_GOOGLE = False


SCOPES = [
    "https://www.googleapis.com/auth/drive",
    "https://www.googleapis.com/auth/spreadsheets",
]

# Header names that map to pay and receipt columns (matches excel_parser.py HEADER_MAP)
_PAY_HEADERS = {"pay", "fees", "amount"}
_RECEIPT_HEADERS = {"receipt no.", "receipt num", "receipt number", "receipt no"}


def _get_credentials():
    """Build Google credentials from settings."""
    import os, json, base64

    creds_value = settings.google_drive_credentials_json.strip()
    if not creds_value:
        raise RuntimeError("GOOGLE_DRIVE_CREDENTIALS_JSON not configured")

    if os.path.isfile(creds_value):
        return service_account.Credentials.from_service_account_file(creds_value, scopes=SCOPES)

    if not creds_value.startswith("{"):
        try:
            creds_value = base64.b64decode(creds_value).decode("utf-8")
        except Exception:
            pass
    try:
        info = json.loads(creds_value)
    except json.JSONDecodeError:
        info = json.loads(creds_value, strict=False)

    return service_account.Credentials.from_service_account_info(info, scopes=SCOPES)


def _col_letter(col_idx: int) -> str:
    """Convert 1-based column index to Excel column letter (1→A, 2→B, 27→AA)."""
    result = ""
    while col_idx > 0:
        col_idx, remainder = divmod(col_idx - 1, 26)
        result = chr(65 + remainder) + result
    return result


def update_athlete_payment(
    drive_file_id: str,
    athlete_number: int,
    pay_value: str,
    receipt_number: str,
) -> bool:
    """Write pay amount and receipt number to the athlete's row in the Google Drive file.

    Finds the roster sheet (tab with 'Reg' in name), locates the athlete by ID,
    and updates the Pay and Receipt columns.

    Uses a per-file lock so concurrent payments to the same branch are serialized,
    preventing the download-modify-upload race condition.

    Returns True if successful. Raises on error so caller can capture the message.
    """
    if not _HAS_GOOGLE:
        raise RuntimeError("Google libraries not installed — cannot write to Drive")

    lock = _get_file_lock(drive_file_id)
    # Acquire lock — concurrent writes to the same file wait here
    with lock:
        return _retry(_do_update, drive_file_id, athlete_number, pay_value, receipt_number)


def _do_update(drive_file_id: str, athlete_number: int, pay_value: str, receipt_number: str) -> bool:
    """Internal: perform the actual Drive write (called under lock with retry)."""
    creds = _get_credentials()

    # Check file type
    drive_service = build("drive", "v3", credentials=creds, cache_discovery=False)
    file_meta = drive_service.files().get(fileId=drive_file_id, fields="mimeType").execute()
    mime = file_meta.get("mimeType", "")

    if mime == "application/vnd.google-apps.spreadsheet":
        return _update_google_sheet(creds, drive_file_id, athlete_number, pay_value, receipt_number)
    else:
        return _update_xlsx_on_drive(creds, drive_service, drive_file_id, athlete_number, pay_value, receipt_number)


def _find_roster_sheet_name(sheets_service, spreadsheet_id: str) -> Optional[str]:
    """Find the roster sheet name (contains 'Reg') in a Google Sheet."""
    meta = sheets_service.spreadsheets().get(
        spreadsheetId=spreadsheet_id, fields="sheets.properties.title"
    ).execute()
    for sheet in meta.get("sheets", []):
        title = sheet["properties"]["title"]
        if "reg" in title.lower():
            return title
    return None


def read_existing_receipts(drive_file_id: str) -> list[str]:
    """Read all existing receipt numbers from the roster sheet.

    Uses the same per-file lock to prevent reading during a concurrent write.
    """
    if not _HAS_GOOGLE:
        return []

    lock = _get_file_lock(drive_file_id)
    with lock:
        return _read_existing_receipts_unlocked(drive_file_id)


def _read_existing_receipts_unlocked(drive_file_id: str) -> list[str]:
    """Internal: read receipts without acquiring lock (caller holds it)."""
    creds = _get_credentials()
    sheets = build("sheets", "v4", credentials=creds, cache_discovery=False)
    sheet_name = _find_roster_sheet_name(sheets, drive_file_id)
    if not sheet_name:
        return []

    result = sheets.spreadsheets().values().get(
        spreadsheetId=drive_file_id,
        range=f"'{sheet_name}'!A1:Z2000",
        valueRenderOption="UNFORMATTED_VALUE",
    ).execute()
    rows = result.get("values", [])

    # Find header row and receipt column
    receipt_col = None
    header_row_idx = None
    for i, row in enumerate(rows):
        has_id = False
        has_name = False
        for j, cell in enumerate(row):
            cell_clean = str(cell).strip().lower()
            if cell_clean in ("id", "no", "no."):
                has_id = True
            elif cell_clean == "name":
                has_name = True
            elif cell_clean in _RECEIPT_HEADERS:
                receipt_col = j
        if has_id and has_name and receipt_col is not None:
            header_row_idx = i
            break

    if receipt_col is None or header_row_idx is None:
        return []

    receipts = []
    for i in range(header_row_idx + 1, len(rows)):
        row = rows[i]
        if receipt_col < len(row):
            val = str(row[receipt_col]).strip()
            if val and val != "None":
                receipts.append(val)
    return receipts


def _update_google_sheet(
    creds, spreadsheet_id: str, athlete_number: int,
    pay_value: str, receipt_number: str,
) -> bool:
    """Update cells in a native Google Sheet using the Sheets API."""
    sheets = build("sheets", "v4", credentials=creds, cache_discovery=False)

    # Find roster sheet
    sheet_name = _find_roster_sheet_name(sheets, spreadsheet_id)
    if not sheet_name:
        raise RuntimeError(f"No roster sheet ('Reg') found in spreadsheet {spreadsheet_id}")

    # Read all data to find header row and athlete row
    result = sheets.spreadsheets().values().get(
        spreadsheetId=spreadsheet_id,
        range=f"'{sheet_name}'!A1:Z2000",
        valueRenderOption="UNFORMATTED_VALUE",
    ).execute()
    rows = result.get("values", [])

    if not rows:
        raise RuntimeError("Roster sheet is empty")

    # Find header row (first row with 'id' and 'name' columns)
    header_row_idx = None
    col_map = {}
    for i, row in enumerate(rows):
        matches = {}
        for j, cell in enumerate(row):
            cell_clean = str(cell).strip().lower()
            if cell_clean in ("id", "no", "no."):
                matches["id"] = j
            elif cell_clean == "name":
                matches["name"] = j
            elif cell_clean in _PAY_HEADERS:
                matches["pay"] = j
            elif cell_clean in _RECEIPT_HEADERS:
                matches["receipt_no"] = j
        if "id" in matches and "name" in matches:
            header_row_idx = i
            col_map = matches
            break

    if header_row_idx is None:
        raise RuntimeError("Could not find header row with ID and Name columns")

    id_col = col_map["id"]
    pay_col = col_map.get("pay")
    receipt_col = col_map.get("receipt_no")

    if pay_col is None and receipt_col is None:
        raise RuntimeError("No Pay or Receipt column found in roster sheet")

    # Find the athlete's row by athlete number
    athlete_row_idx = None
    for i in range(header_row_idx + 1, len(rows)):
        row = rows[i]
        if id_col >= len(row):
            continue
        cell_val = str(row[id_col]).strip()
        # Parse athlete number — strip prefix letters (M, C, E, etc.)
        num_match = re.search(r"(\d+)", cell_val)
        if num_match and int(num_match.group(1)) == athlete_number:
            athlete_row_idx = i
            break

    if athlete_row_idx is None:
        raise RuntimeError(f"Athlete #{athlete_number} not found in roster sheet")

    # Sheets API uses 1-based rows, 0-based columns for A1 notation
    sheet_row = athlete_row_idx + 1  # Convert to 1-based

    # Build batch update
    updates = []
    if pay_col is not None:
        pay_cell = f"'{sheet_name}'!{_col_letter(pay_col + 1)}{sheet_row}"
        updates.append({"range": pay_cell, "values": [[pay_value]]})
    if receipt_col is not None:
        receipt_cell = f"'{sheet_name}'!{_col_letter(receipt_col + 1)}{sheet_row}"
        updates.append({"range": receipt_cell, "values": [[receipt_number]]})

    if updates:
        sheets.spreadsheets().values().batchUpdate(
            spreadsheetId=spreadsheet_id,
            body={"valueInputOption": "USER_ENTERED", "data": updates},
        ).execute()
        logger.info(
            f"Updated athlete #{athlete_number} in sheet: Pay={pay_value}, Receipt={receipt_number}"
        )

    return True


def _update_xlsx_on_drive(
    creds, drive_service, file_id: str, athlete_number: int,
    pay_value: str, receipt_number: str,
) -> bool:
    """Download .xlsx, modify with openpyxl, re-upload to Drive."""
    import io
    import tempfile
    import os
    from openpyxl import load_workbook
    from googleapiclient.http import MediaIoBaseDownload, MediaFileUpload

    # Download
    request = drive_service.files().get_media(fileId=file_id)
    tmp_path = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False).name
    with io.FileIO(tmp_path, "wb") as f:
        downloader = MediaIoBaseDownload(f, request)
        done = False
        while not done:
            _, done = downloader.next_chunk()

    try:
        wb = load_workbook(tmp_path)

        # Find roster sheet
        roster_ws = None
        for name in wb.sheetnames:
            if "reg" in name.lower():
                roster_ws = wb[name]
                break

        if not roster_ws:
            logger.error("No roster sheet found in .xlsx")
            return False

        # Find header row and columns
        header_row = None
        pay_col = None
        receipt_col = None
        id_col = None

        for row_idx in range(1, min(20, (roster_ws.max_row or 0) + 1)):
            has_id = False
            has_name = False
            for col_idx in range(1, min(30, (roster_ws.max_column or 0) + 1)):
                val = str(roster_ws.cell(row=row_idx, column=col_idx).value or "").strip().lower()
                if val in ("id", "no", "no."):
                    id_col = col_idx
                    has_id = True
                elif val == "name":
                    has_name = True
                elif val in _PAY_HEADERS:
                    pay_col = col_idx
                elif val in _RECEIPT_HEADERS:
                    receipt_col = col_idx
            if has_id and has_name:
                header_row = row_idx
                break

        if header_row is None or id_col is None:
            logger.error("Could not find header row in .xlsx roster sheet")
            return False

        # Find athlete row
        athlete_row = None
        for row_idx in range(header_row + 1, (roster_ws.max_row or 0) + 1):
            cell_val = str(roster_ws.cell(row=row_idx, column=id_col).value or "").strip()
            num_match = re.search(r"(\d+)", cell_val)
            if num_match and int(num_match.group(1)) == athlete_number:
                athlete_row = row_idx
                break

        if athlete_row is None:
            logger.error(f"Athlete #{athlete_number} not found in .xlsx roster")
            return False

        # Write values
        if pay_col:
            roster_ws.cell(row=athlete_row, column=pay_col, value=pay_value)
        if receipt_col:
            roster_ws.cell(row=athlete_row, column=receipt_col, value=receipt_number)

        wb.save(tmp_path)
        wb.close()

        # Re-upload
        media = MediaFileUpload(
            tmp_path,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        drive_service.files().update(fileId=file_id, media_body=media).execute()

        logger.info(
            f"Updated athlete #{athlete_number} in .xlsx: Pay={pay_value}, Receipt={receipt_number}"
        )
        return True

    finally:
        try:
            os.unlink(tmp_path)
        except OSError:
            pass
