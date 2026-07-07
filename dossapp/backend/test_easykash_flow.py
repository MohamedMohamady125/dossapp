"""End-to-end test of Easykash payment flow (run: venv/bin/python test_easykash_flow.py)."""

import hashlib
import hmac
import os
import sqlite3
from datetime import datetime

TEST_DB = "test_easykash.db"
TEST_WB = "test_easykash_workbook.xlsx"
for f in (TEST_DB, TEST_WB):
    if os.path.exists(f):
        os.remove(f)

os.environ["DATABASE_URL"] = f"sqlite+aiosqlite:///./{TEST_DB}"
os.environ["EASYKASH_API_KEY"] = "test-api-key"
os.environ["EASYKASH_HMAC_SECRET"] = "test-hmac-secret"
os.environ["EASYKASH_PAYMENT_OPTIONS"] = "2,4"
os.environ["EASYKASH_ALLOW_UNVERIFIED"] = "false"
os.environ["DRIVE_FILE_ID_BRANCH_1"] = ""  # don't touch the real Drive workbook
os.environ["VERCEL"] = "1"  # serverless mode: no background refresh/reconciliation task

from openpyxl import Workbook  # noqa: E402


def make_workbook(path: str):
    wb = Workbook()
    reg = wb.active
    reg.title = "Reg."
    reg.append(["ID", "Name", "Pay", "Phone 1"])
    reg.append(["M1", "Omar Hassan", None, "01001234567"])  # unpaid → can pay online
    att = wb.create_sheet("Sat.Wed")
    att.cell(row=2, column=1, value="Coach")
    att.cell(row=2, column=2, value="ID")
    att.cell(row=2, column=3, value="Name")
    att.cell(row=2, column=4, value="Pay")
    att.cell(row=3, column=1, value="5:15 to 6:10 pm")
    att.cell(row=4, column=1, value="Ahmed")
    att.cell(row=4, column=2, value="M1")
    att.cell(row=4, column=3, value="Omar Hassan")
    # NOTE: no Pay value anywhere — a filled Pay cell means "paid cash" (reconciled).
    wb.save(path)


make_workbook(TEST_WB)

from fastapi.testclient import TestClient  # noqa: E402
import app.main as main_mod  # noqa: E402
from app.config import settings  # noqa: E402
from app.services import easykash  # noqa: E402
from app.services.excel_roster_source import ExcelRosterSource  # noqa: E402
from app.utils.auth import create_access_token  # noqa: E402

passed, failed = 0, 0


def check(label, cond, extra=""):
    global passed, failed
    if cond:
        passed += 1
        print(f"  PASS: {label}")
    else:
        failed += 1
        print(f"  FAIL: {label} {extra}")


def sign(payload: dict) -> str:
    fields = [f.strip() for f in settings.easykash_signature_fields.split(",")]
    concatenated = "".join(str(payload.get(f, "")) for f in fields)
    return hmac.new(b"test-hmac-secret", concatenated.encode(), hashlib.sha256).hexdigest()


PERIOD = datetime.now().strftime("%Y-%m")
_now = datetime.now()
PREV_PERIOD = f"{_now.year - 1}-12" if _now.month == 1 else f"{_now.year}-{_now.month - 1:02d}"


def callback_payload(**overrides):
    payload = {
        "Amount": "800",
        "Currency": "EGP",
        "PaymentMethod": "Card",
        "easykashRef": "EK-TEST-001",
        "customerReference": f"AQUA-1-1-{PERIOD}",
        "status": "PAID",
    }
    payload.update(overrides)
    payload["signatureHash"] = sign(payload)
    return payload


# Mock httpx for checkout creation
class MockResponse:
    status_code = 200
    text = ""

    def json(self):
        return {"redirectUrl": "https://easykash.net/pay/checkout/TEST123"}


class MockAsyncClient:
    captured = {}

    def __init__(self, **kwargs):
        pass

    async def __aenter__(self):
        return self

    async def __aexit__(self, *args):
        return False

    async def post(self, url, json=None, headers=None):
        MockAsyncClient.captured = {"url": url, "json": json, "headers": headers}
        return MockResponse()


easykash.httpx.AsyncClient = MockAsyncClient

with TestClient(main_mod.app) as client:
    main_mod.roster_source = ExcelRosterSource([
        {"branch_id": 1, "branch_name": "Rehab", "local_file_path": TEST_WB},
    ])

    conn = sqlite3.connect(TEST_DB)
    conn.execute(
        "INSERT INTO accounts (branch_id, athlete_number, login_code, password_hash, must_change_password,"
        " name_at_creation, status, created_by_admin_id, created_at)"
        " VALUES (1, 1, 'REH-1', 'x', 0, 'Omar Hassan', 'active', 1, CURRENT_TIMESTAMP)"
    )
    # Prior-month cash payment → becomes this month's bill amount (Excel Pay empty = unpaid)
    conn.execute(
        "INSERT INTO payments (branch_id, athlete_number, period, source, amount_paid, currency,"
        " status, paid_at, created_at)"
        f" VALUES (1, 1, '{PREV_PERIOD}', 'cash', 800, 'EGP', 'paid', CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)"
    )
    conn.commit()
    omar_id = conn.execute("SELECT id FROM accounts WHERE athlete_number=1").fetchone()[0]
    conn.close()

    omar_token = create_access_token({"sub": str(omar_id), "role": "customer", "type": "access"})
    omar_h = {"Authorization": f"Bearer {omar_token}"}

    print("\n[1] Checkout creation")
    bill = client.get("/me/bill", headers=omar_h).json()
    check("bill amount from last month's payment", bill.get("amount_owed") == "800" and bill.get("is_paid") is False,
          str(bill))
    r = client.post("/me/pay/easykash/checkout", headers=omar_h)
    check("checkout 200", r.status_code == 200, r.text)
    body = r.json() if r.status_code == 200 else {}
    check("checkout returns url", body.get("url") == "https://easykash.net/pay/checkout/TEST123", str(body))
    check("amount is 800", body.get("amount") == "800", str(body))
    req = MockAsyncClient.captured
    check("API key in Authorization header", req.get("headers", {}).get("Authorization") == "test-api-key")
    check("customerReference format", req.get("json", {}).get("customerReference") == f"AQUA-1-1-{PERIOD}",
          str(req.get("json", {}).get("customerReference")))
    check("paymentOptions from config", req.get("json", {}).get("paymentOptions") == [2, 4],
          str(req.get("json", {}).get("paymentOptions")))
    check("amount in EGP units (not cents)", req.get("json", {}).get("amount") == "800")

    print("\n[2] Webhook: bad signature rejected")
    bad = callback_payload()
    bad["signatureHash"] = "0" * 64
    r = client.post("/webhooks/easykash", json=bad)
    check("403 on bad signature", r.status_code == 403, f"got {r.status_code}")

    print("\n[3] Webhook: valid signature records payment + receipt")
    r = client.post("/webhooks/easykash", json=callback_payload())
    check("webhook 200", r.status_code == 200, r.text)
    check("receipt issued", r.json().get("status") == "ok" and r.json().get("receipt", "").startswith("P-"), r.text)

    receipts = client.get("/me/receipts", headers=omar_h).json()
    check("receipt visible to customer", len(receipts) == 1, str(receipts))

    bill = client.get("/me/bill", headers=omar_h).json()
    check("bill shows paid", bill.get("is_paid") is True, str(bill))

    print("\n[4] Webhook: duplicate callback ignored")
    r = client.post("/webhooks/easykash", json=callback_payload())
    check("duplicate ignored", r.json().get("status") == "duplicate", r.text)
    receipts = client.get("/me/receipts", headers=omar_h).json()
    check("still one receipt", len(receipts) == 1, str(len(receipts)))

    print("\n[5] Webhook: non-success status ignored")
    r = client.post("/webhooks/easykash", json=callback_payload(customerReference=f"AQUA-1-1-2099-01", status="FAILED"))
    check("failed status ignored", r.json().get("status") == "ignored", r.text)

    print("\n[6] Webhook: GET callback with query params works")
    payload = callback_payload(easykashRef="EK-TEST-002", customerReference="AQUA-1-1-2099-02")
    r = client.get("/webhooks/easykash", params=payload)
    check("GET callback 200 + recorded", r.status_code == 200 and r.json().get("status") == "ok", r.text)

    print("\n[7] Already-paid guard on checkout")
    r = client.post("/me/pay/easykash/checkout", headers=omar_h)
    check("checkout blocked after payment", r.status_code == 400, f"got {r.status_code}: {r.text}")

    print("\n[8] Success landing page")
    r = client.get("/pay/success")
    check("success page 200 HTML", r.status_code == 200 and "Payment Received" in r.text)

    print("\n[9] Unconfigured HMAC secret rejects callbacks (allow_unverified=false)")
    old = settings.easykash_hmac_secret
    settings.easykash_hmac_secret = ""
    r = client.post("/webhooks/easykash", json=callback_payload(easykashRef="EK-TEST-003"))
    check("503 when secret missing", r.status_code == 503, f"got {r.status_code}")
    settings.easykash_allow_unverified = True
    r = client.post("/webhooks/easykash", json=callback_payload(customerReference="AQUA-1-1-2099-03"))
    check("accepted when allow_unverified", r.status_code == 200 and r.json().get("status") == "ok", r.text)
    settings.easykash_hmac_secret = old
    settings.easykash_allow_unverified = False

    print("\n[10] Checkout 503 when API key unset")
    settings.easykash_api_key = ""
    r = client.post("/me/pay/easykash/checkout", headers=omar_h)
    check("503 without API key", r.status_code == 503, f"got {r.status_code}")
    settings.easykash_api_key = "test-api-key"

print(f"\n{'='*40}\n{passed} passed, {failed} failed")
for f in (TEST_DB, TEST_WB):
    if os.path.exists(f):
        os.remove(f)
exit(1 if failed else 0)
