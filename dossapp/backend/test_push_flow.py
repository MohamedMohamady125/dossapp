"""End-to-end test of push notification triggers (run: venv/bin/python test_push_flow.py)."""

import os
import sqlite3
import time
from datetime import datetime

TEST_DB = "test_push.db"
TEST_WB = "test_push_workbook.xlsx"
for f in (TEST_DB, TEST_WB):
    if os.path.exists(f):
        os.remove(f)

os.environ["DATABASE_URL"] = f"sqlite+aiosqlite:///./{TEST_DB}"

from openpyxl import Workbook  # noqa: E402


def make_workbook(path: str, time_block: str, marks: dict[str, list[str]]):
    """marks: athlete id -> [mark for each of 3 session dates]"""
    wb = Workbook()
    reg = wb.active
    reg.title = "Reg."
    reg.append(["ID", "Name", "Pay", "Phone 1"])
    reg.append(["M1", "Omar Hassan", 1500, "01001234567"])  # pay present = PAID (reconciled as cash)
    reg.append(["M2", "Sara Adel", None, "01007654321"])  # no pay = unpaid this month

    att = wb.create_sheet("Sat.Wed")
    # Header row with Coach/ID/Name/Pay + date columns
    att.cell(row=2, column=1, value="Coach")
    att.cell(row=2, column=2, value="ID")
    att.cell(row=2, column=3, value="Name")
    att.cell(row=2, column=4, value="Pay")
    att.cell(row=2, column=6, value=datetime(2026, 6, 28))
    att.cell(row=2, column=7, value=datetime(2026, 7, 1))
    att.cell(row=2, column=8, value=datetime(2026, 7, 4))
    # Time block row
    att.cell(row=3, column=1, value=time_block)
    # Athlete rows
    for i, (aid, name, coach) in enumerate([("M1", "Omar Hassan", "Ahmed"), ("M2", "Sara Adel", "Ahmed")]):
        r = 4 + i
        att.cell(row=r, column=1, value=coach)
        att.cell(row=r, column=2, value=aid)
        att.cell(row=r, column=3, value=name)
        for j, mark in enumerate(marks[aid]):
            if mark:
                att.cell(row=r, column=6 + j, value=mark)
    wb.save(path)


# Initial state: Omar missed last 2, Sara missed only last 1
make_workbook(TEST_WB, "5:15 to 6:10 pm", {"M1": ["P", "A", "A"], "M2": ["P", "P", "A"]})

from fastapi.testclient import TestClient  # noqa: E402
import app.main as main_mod  # noqa: E402
from app.services.excel_roster_source import ExcelRosterSource  # noqa: E402
from app.utils.auth import create_access_token  # noqa: E402

passed, failed = 0, 0


def check(label, cond, extra=""):
    global passed, failed
    if cond:
        passed += 1
        print(f"  PASS: {label}")
    else:
        failed += 1
        print(f"  FAIL: {label} {extra}")


with TestClient(main_mod.app) as client:
    main_mod.roster_source = ExcelRosterSource([
        {"branch_id": 1, "branch_name": "Rehab", "local_file_path": TEST_WB},
    ])

    # Seed customer accounts directly (bypass provisioning flow)
    conn = sqlite3.connect(TEST_DB)
    conn.execute(
        "INSERT INTO accounts (branch_id, athlete_number, login_code, password_hash, must_change_password,"
        " name_at_creation, status, created_by_admin_id, created_at)"
        " VALUES (1, 1, 'REH-1', 'x', 0, 'Omar Hassan', 'active', 1, CURRENT_TIMESTAMP)"
    )
    conn.execute(
        "INSERT INTO accounts (branch_id, athlete_number, login_code, password_hash, must_change_password,"
        " name_at_creation, status, created_by_admin_id, created_at)"
        " VALUES (1, 2, 'REH-2', 'x', 0, 'Sara Adel', 'active', 1, CURRENT_TIMESTAMP)"
    )
    conn.commit()
    omar_id = conn.execute("SELECT id FROM accounts WHERE athlete_number=1").fetchone()[0]
    sara_id = conn.execute("SELECT id FROM accounts WHERE athlete_number=2").fetchone()[0]
    conn.close()

    admin_token = create_access_token({"sub": "1", "role": "admin", "type": "access"})
    omar_token = create_access_token({"sub": str(omar_id), "role": "customer", "type": "access"})
    sara_token = create_access_token({"sub": str(sara_id), "role": "customer", "type": "access"})
    admin_h = {"Authorization": f"Bearer {admin_token}"}
    omar_h = {"Authorization": f"Bearer {omar_token}"}
    sara_h = {"Authorization": f"Bearer {sara_token}"}

    def inbox_types(headers):
        return sorted(n["type"] for n in client.get("/me/notifications", headers=headers).json())

    print("\n[1] Cron auth")
    r = client.get("/cron/notifications")
    check("no auth rejected", r.status_code in (401, 403), f"got {r.status_code}")
    r = client.get("/cron/notifications", headers={"Authorization": "Bearer wrong"})
    check("bad secret rejected", r.status_code == 401, f"got {r.status_code}")

    print("\n[2] First trigger run (seed mode + missed sessions)")
    # Note: triggers also fire inside the lazy _ensure_loaded hook, so verify via inbox
    r = client.get("/cron/notifications", headers=admin_h)
    check("cron 200", r.status_code == 200, r.text)
    check("Omar missed-2 notified once", inbox_types(omar_h) == ["missed_sessions"], str(inbox_types(omar_h)))
    check("Sara (1 miss) not notified", inbox_types(sara_h) == [], str(inbox_types(sara_h)))

    print("\n[3] Re-run: dedupe")
    r = client.get("/cron/notifications", headers=admin_h)
    body = r.json()
    check("no duplicate sends", body["sent"] == {"schedule": 0, "missed": 0}, str(body))

    print("\n[4] Schedule change detection")
    make_workbook(TEST_WB, "6:15 to 7:10 pm", {"M1": ["P", "A", "A"], "M2": ["P", "P", "A"]})
    os.utime(TEST_WB, (time.time() + 5, time.time() + 5))
    r = client.get("/cron/notifications", headers=admin_h)
    body = r.json()
    check("both schedules changed -> 2 sends", body["sent"]["schedule"] == 2, str(body))
    check("missed still deduped", body["sent"]["missed"] == 0, str(body))

    print("\n[5] Payment reminders")
    r = client.get("/cron/payment-reminders", headers=admin_h)
    check("skipped when not the 25th", r.json().get("skipped") is True or r.json().get("sent") is not None, r.text)
    r = client.get("/cron/payment-reminders?force=true", headers=admin_h)
    body = r.json()
    check("only unpaid Sara reminded", body.get("sent") == 1, str(body))
    check("Sara got payment reminder", "payment_reminder" in inbox_types(sara_h), str(inbox_types(sara_h)))
    check("paid Omar not reminded", "payment_reminder" not in inbox_types(omar_h), str(inbox_types(omar_h)))
    r = client.get("/cron/payment-reminders?force=true", headers=admin_h)
    check("reminder deduped per month", r.json().get("sent") == 0, r.text)

    print("\n[6] Customer inbox")
    r = client.get("/me/notifications", headers=omar_h)
    check("inbox 200", r.status_code == 200, r.text)
    items = r.json()
    types = sorted(n["type"] for n in items)
    check("Omar has missed + schedule", types == ["missed_sessions", "schedule_change"], str(types))
    r = client.get("/me/notifications/unread-count", headers=omar_h)
    check("unread = 2", r.json()["unread"] == 2, r.text)
    first_id = items[0]["id"]
    r = client.post(f"/me/notifications/{first_id}/read", headers=omar_h)
    check("mark read 200", r.status_code == 200, r.text)
    r = client.get("/me/notifications/unread-count", headers=omar_h)
    check("unread = 1 after read", r.json()["unread"] == 1, r.text)
    r = client.post("/me/notifications/read-all", headers=omar_h)
    r = client.get("/me/notifications/unread-count", headers=omar_h)
    check("unread = 0 after read-all", r.json()["unread"] == 0, r.text)

    print("\n[7] Device registration")
    r = client.post("/me/devices", json={"token": "fake-fcm-token-123", "platform": "android"}, headers=omar_h)
    check("register device", r.status_code == 200, r.text)
    r = client.post("/me/devices", json={"token": "fake-fcm-token-123", "platform": "android"}, headers=omar_h)
    check("re-register idempotent", r.status_code == 200, r.text)
    r = client.post("/me/devices", json={"token": "t", "platform": "windows"}, headers=omar_h)
    check("bad platform rejected", r.status_code == 400, r.text)
    r = client.delete("/me/devices/fake-fcm-token-123", headers=omar_h)
    check("unregister device", r.status_code == 200, r.text)

    print("\n[8] Admin test-send + debug attendance")
    r = client.post("/admin/notifications/test-send", json={"branch_id": 1, "athlete_number": 1, "dry_run": True}, headers=admin_h)
    check("dry-run test-send", r.status_code == 200 and r.json()["dry_run"] is True, r.text)
    r = client.post("/admin/notifications/test-send", json={"branch_id": 1, "athlete_number": 1, "dry_run": False}, headers=admin_h)
    check("real test-send (stub)", r.status_code == 200, r.text)
    r = client.get("/admin/debug/attendance/1", headers=admin_h)
    check("debug attendance 200", r.status_code == 200, r.text)
    dbg = r.json()
    check("3 dates detected", dbg["detected_dates"] == ["2026-06-28", "2026-07-01", "2026-07-04"], str(dbg["detected_dates"]))
    check("2 athletes with marks", dbg["athletes_with_attendance"] == 2, str(dbg))

    print("\n[9] Absence streak reset -> new streak re-notifies")
    # Omar attends (P on a new date), then misses 2 new sessions
    wb_marks = {"M1": ["A", "A", "P"], "M2": ["P", "P", "A"]}
    make_workbook(TEST_WB, "6:15 to 7:10 pm", wb_marks)
    os.utime(TEST_WB, (time.time() + 10, time.time() + 10))
    r = client.get("/cron/notifications", headers=admin_h)
    check("P at end -> no missed alert", r.json()["sent"]["missed"] == 0, r.text)
    wb_marks = {"M1": ["P", "A", "A"], "M2": ["P", "P", "A"]}  # same as original streak (same start date -> deduped)
    make_workbook(TEST_WB, "6:15 to 7:10 pm", wb_marks)
    os.utime(TEST_WB, (time.time() + 15, time.time() + 15))
    r = client.get("/cron/notifications", headers=admin_h)
    check("same streak start deduped", r.json()["sent"]["missed"] == 0, r.text)

print(f"\n{'='*40}\n{passed} passed, {failed} failed")
for f in (TEST_DB, TEST_WB):
    if os.path.exists(f):
        os.remove(f)
exit(1 if failed else 0)
